import openpyxl

"""
All functions for reading test data in excel file.
"""

path = ".//TestData//LoginData.xlsx"
"""
    Gets total row count of test data
    """
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    return sheet.max_row

def getColumnCount(file, sheetName):
    """
    Gets total column count of test data
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    return sheet.max_column

def readData(file, sheetName, rownum, colnum):
    """
    Reads value of cell at given row and column
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    return sheet.cell(row=rownum, column=colnum).value


def testData(file, sheetName):
    """
    Returns the Test data for the specific scenario and skips all data that has execution set to no.
    """

    rows = getRowCount(path, 'Sheet1')
    cols = getRowCount(path, 'Sheet1')


    for r in range(2, rows + 1):
        if readData(path, 'Sheet1', r, 2).lower() == 'no':
            continue

        data = []
        data.append(readData(path, 'Sheet1', r, 1))
        data.append(readData(path, 'Sheet1', r, 2))
        data.append(readData(path, 'Sheet1', r, 3))
        data.append(readData(path, 'Sheet1', r, 4))
        data.append(readData(path, 'Sheet1', r, 5))
        data.append(readData(path, 'Sheet1', r, 6))
        data.append(readData(path, 'Sheet1', r, 7))
        data.append(readData(path, 'Sheet1', r, 8))
        data.append(readData(path, 'Sheet1', r, 9))
        data.append(readData(path, 'Sheet1', r, 10))

        print(data)

        yield data




